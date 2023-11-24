import openpyxl
import selenium
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import time
from openpyxl import Workbook, load_workbook
import numpy as np
import pandas as pd

service = Service()
option = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=option)

name = []
with open("people.csv", "r") as file:         
    next(file)                                  # sak ar otro liniju
    for line in file:                            # iet cauri katrai rindai  
        row = line.rstrip().split(",")            # nonem visas space bars pec varda, katru vardu atdala ar komatu 
        name.append(row[2] + ' ' + row[3])        # append pievieno sarakstam ( name ) 2 un 3 kollonu

url = "https://emn178.github.io/online-tools/crc32.html"        
driver.get(url)
time.sleep(2)                                              # pagaida 2s lai kaut ko izdaritu

wb = load_workbook('salary.xlsx')                        # loado, jeb atlauj kaut ko darit ar excel failu
ws = wb.active                                              # ļāuj nolasit
max_row = ws.max_row                                            # atrod galējo rindu
sheet = wb["result"]                                              # mainigajam sheet pieder excel lapa "result" 
j = 1
for x in name:
    find = driver.find_element(By.ID, "input")                          # atrod vietu kur ierakstit
    find.send_keys(x)                                                   # aizsuta x, kur x ir katrs vards no sarakasta name       
    find = driver.find_element(By.ID, "output")                           # atrod vietu kur ir rezultats
    temp = find.get_attribute("value")                                    # nolasa, kas no ailites

    rezultats = 0
    for i in range(1, max_row + 1):                                      # no pirmas lidz pedejai rindai
        cilvekaKods = ws['A' + str(i)].value                              # nolasa 'A' kollonas šunas vertibu
        alga = ws['B' + str(i)].value                                     # nolasa 'B' kollonas šunas vertibu 
        if cilvekaKods == temp:                                           # ja 'A' šunas vertiba ir vienada ar kodu, kurs tika nolasits no interneta
            rezultats = rezultats + alga                                  # tad to pieskaita
                                                                            # !tas atkartojas ar katru cilveku un vina personigo kodu
    sheet['A' + str(j)].value = x                               # "result" excel lapa ; 'A' kollona liek vertibas x ( cilveka vardus )
    sheet['C' + str(j)].value = rezultats                       #  "result" excel lapa ; 'C' kollona liek vertibas - galejo algu 
    wb.save('salary.xlsx')                                       # visu informaciju saglaba
    j = j + 1                                                    # vajag paprasit

    find = driver.find_element(By.ID, "input")                    # atrod vietu kur ierakstit
    find.click()                                                   # uzklišķina
    driver.refresh()                                                # refresho majaslapu
